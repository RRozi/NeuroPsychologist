import json

import shuttleai

from db import connect, cursor
import time
import logging
from params import session
import openpyxl
from shuttleai import *
import os
# Suttle
shuttle = ShuttleClient(api_key=os.getenv("SHUTTLE"))

# Работа с Exls
if not os.path.exists("sessions.xlsx"):
    wb = openpyxl.Workbook()
    wb.save("sessions.xlsx")

workbook = openpyxl.load_workbook("sessions.xlsx")
sheet = workbook.active

# function write info in history
def LogInExls(idCol: int, value: list):
    correct_value = ''
    for x in value:
        correct_value += '\n'
        for k, v in x.items():
            correct_value += f"{k}: {v}\n"

    if session.EXLS_LAST_ROW_ID == None:
        session.EXLS_LAST_ROW_ID = sheet.max_row + 1

        sheet.cell(
            row=session.EXLS_LAST_ROW_ID,
            column=1,
            value=session.USER_SESSION_ID) # Запись id сессии в активную строку exls сессии

    sheet.cell(
        row=session.EXLS_LAST_ROW_ID,
        column=idCol,
        value=correct_value)

    workbook.save("sessions.xlsx")


def add_History(role: str, content: str) -> None:
    session.HISTORY_DATEBASE.append(
        {
            "role": role,
            "content": content
        }
    )
    session.HISTORY.append(
        {
            "role":role,
            "content":content
        }
    )

def request_(UserPromt: str, time_out = 1) -> str:
    try:

        add_History("user", UserPromt)  # Запись ответа ПОЛЬЗОВАТЕЛЯ в ИСТОРИЮ
        request = shuttle.chat_completion(
            model=session.GPT_MODEL,
            messages=session.HISTORY,
            stream=False,
        )['choices'][0]['message']['content']
        add_History("assistant", request)  # Запись ответа GPT в ИСТОРИЮ
        if session.USER_SESSION_ID:
            cursor.execute("UPDATE sessions SET history = ? WHERE id = ?",
                           (json.dumps(session.HISTORY_DATEBASE, indent=4, ensure_ascii=False), session.USER_SESSION_ID))
            connect.commit()

        else:
            cursor.execute("INSERT INTO sessions (history) VALUES (?)",
                           (json.dumps(session.HISTORY_DATEBASE, indent=4, ensure_ascii=False),))
            connect.commit()
            session.USER_SESSION_ID = cursor.lastrowid
        LogInExls(idCol=2, value=session.HISTORY_DATEBASE)

    except KeyError as e:
        print(e)
        logging.error(f"Ошибка при попытке ответить на вопрос! \n Сессия ---> {session.HISTORY}")
        time.sleep(time_out)
        request_(UserPromt, time_out + 3)

    else:
        print("Нейропсихолог:", request)
        logging.info(f"Ответ дан корректно.")
        return request